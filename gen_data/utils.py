import json
from collections import defaultdict
from name import *
import random
import os
import re
import copy
import sqlite3
import pandas as pd
from collections import OrderedDict
from pyecharts import Bar

random.seed(133)

def mkdir(path):
    if os.path.exists(path):
        print(path + " file exists!")
    else:
        os.mkdir(path)

def read_json(path):
    f = open(path, "r", encoding='utf-8')
    content = json.load(f, object_pairs_hook=OrderedDict)
    f.close()
    return content

def read_jsonl(path):
    f = open(path, "r", encoding='utf-8')
    lines = f.readlines()
    f.close()
    content = []
    for line in lines:
        tmp = eval(line.replace("true", "True").replace("false", "False"))
        content.append(tmp)
    return content

def get_wikitable_question(path):
    f = open(path, "r", encoding='utf-8')
    lines = f.readlines()
    f.close()
    content = []
    for i, line in enumerate(lines):
        if i == 0:
            continue
        tmp = line.split("\t")
        content.append({Question: tmp[1], Wikitabledbid: tmp[2]})
    return content

def get_wikitable_table_data():
    wikitablecsv = os.path.join(Wikitabledir, Csv)
    csvdirs = os.listdir(wikitablecsv)
    wikitable_table_data = {}
    for csvdir in csvdirs:
        wikitablecsvdirpath = os.path.join(wikitablecsv, csvdir)
        csvfiles = os.listdir(wikitablecsvdirpath)
        for csvfile in csvfiles:
            if csvfile.split('.')[1] == 'csv':
                csvpath = os.path.join(wikitablecsvdirpath, csvfile)
                f = open(csvpath, "r", encoding="utf-8")
                header = f.readline()
                f.close()
                wikitable_table_data[Csv + '/' + csvdir + '/' + csvfile] = list(header.strip().lower().replace("\"", "").split(","))

    return wikitable_table_data

def get_hybridQA_table_data():
    tablefiles = os.listdir(HybridQAtablepath)
    hybridQA_table_data = defaultdict(list)
    for tablefile in tablefiles:
        tablefilepath = os.path.join(HybridQAtablepath, tablefile)
        content = read_json(tablefilepath)
        dbid = content[HybridQAtableid]
        header = content[HybridQAheader]
        for item in header:
            column = item[0][0].lower()
            hybridQA_table_data[dbid].append(column)
    return hybridQA_table_data

def get_tablefact_table_data():
    tablefiles = os.listdir(Tablefacttablepath)
    tablefact_table_data = defaultdict(list)
    for tablefile in tablefiles:
        tablefilepath = os.path.join(Tablefacttablepath, tablefile)
        f = open(tablefilepath, "r", encoding="utf-8")
        title = f.readline()
        f.close()
        column = title.strip().lower().split("#")
        tablefact_table_data[tablefile] = column
    return tablefact_table_data

def get_msmarco_usefulness(path):
    f = open(path, "r", encoding='utf-8')
    lines = f.readlines()
    f.close()
    questions = set()
    for i, line in enumerate(lines):
        if i == 0:
            continue
        tmp = line.split("\t")
        questions.add(tmp[1].lower())
    return questions

def get_wikiQAtsv(path):
    f = open(path, "r", encoding='utf-8')
    lines = f.readlines()
    f.close()
    questions = set()
    for i, line in enumerate(lines):
        if i == 0:
            continue
        tmp = line.split("\t")
        questions.add(tmp[1].lower())
    return questions

def get_domainsqlite(path1, path2):
    files = os.listdir(path2)
    db2table = defaultdict(dict)
    for file in files:
        filepath = os.path.join(path1, file.split(".")[0] + ".sqlite")
        with sqlite3.connect(filepath) as con:
            c = con.cursor()
            for tables in c.execute("SELECT name FROM sqlite_master WHERE type='table'"):
                for table in tables:
                    df = pd.read_sql_query("SELECT * FROM " + str(table), con=con)
                    db2table[file.split(".")[0].lower()][str(table).lower()] = [column.replace("\"", "").lower() for column in list(df)]
    return db2table

def get_domainrevised(path):
    files = os.listdir(path)
    db2questionsambiguous = defaultdict(list)
    db2questionsnotambiguous = defaultdict(list)
    for file in files:
        filepath = os.path.join(path, file)
        content = read_json(filepath)
        for item in content:
            metafeature = item[Dbdomainmetafeature]
            if metafeature:
                sentences = item[Dbdomainsentences]
                for sentence in sentences:
                    question = sentence[Dbdomainfulltext].lower()
                    db2questionsambiguous[file.split(".")[0].lower()].append(question)
            else:
                sentences = item[Dbdomainsentences]
                for sentence in sentences:
                    question = sentence[Dbdomainfulltext].lower()
                    db2questionsnotambiguous[file.split(".")[0].lower()].append(question)
    return db2questionsambiguous, db2questionsnotambiguous

def preprocess_data(dataset, datadb, dataname):
    questions = defaultdict(list)
    dbschema = defaultdict(dict)

    if dataname == Spider:
        for item in dataset:
            dbid = item[Spiderdbid].lower()
            question = item[Question].lower()
            query = set([tok.lower().replace("_", " ").split(".")[-1] for tok in item[Spiderquerytok]])
            questions[dbid].append((question, Spider, query))
        for item in datadb:
            dbid = item[Spiderdbid].lower()
            table_names = item[Spidertablename]
            table_names = dict(zip(range(len(table_names)), table_names))
            column_names = item[Spidercolumnname]
            table2column = defaultdict(list)
            column_set = set()
            for column in column_names:
                if column[0] in table_names:
                    table2column[table_names[column[0]].lower()].append(column[1].lower())
                column_set.add(column[1].lower())
            dbschema[dbid][Table] = table2column
            dbschema[dbid][Column] = column_set
    elif dataname == Spiderother:
        for item in dataset:
            dbid = item[Spiderdbid].lower()
            question = item[Question].lower()
            query = set([tok.lower().replace("_", " ").split(".")[-1] for tok in item[Spiderquerytok]])
            questions[dbid].append((question, Spiderother, query))
        for item in datadb:
            dbid = item[Spiderdbid].lower()
            table_names = item[Spidertablename]
            table_names = dict(zip(range(len(table_names)), table_names))
            column_names = item[Spidercolumnname]
            table2column = defaultdict(list)
            column_set = set()
            for column in column_names:
                if column[0] in table_names:
                    table2column[table_names[column[0]].lower()].append(column[1].lower())
                column_set.add(column[1].lower())
            dbschema[dbid][Table] = table2column
            dbschema[dbid][Column] = column_set
    elif dataname == HybridQA:
        for item in dataset:
            dbid = str(item[HybridQAdbid]).lower()
            question = item[Question].lower()
            query = set(question.lower().split(" "))
            #query = set()
            questions[dbid].append((question, HybridQA, query))
        for item in datadb:
            table = datadb[item]
            dbschema[str(item).lower()][Table] = {str(item).lower(): table}
            dbschema[str(item).lower()][Column] = set(table)
    elif dataname == WikiSQL:
        for item in datadb:
            dbid = item[WikiSQLtable]
            dbschema[dbid][Table] = {dbid: [name.lower() for name in item[WikiSQLheader]]}
            dbschema[dbid][Column] = set([name.lower() for name in item[WikiSQLheader]])
        for item in dataset:
            dbid = item[WikiSQLdbid].lower()
            question = item[Question].lower()
            sql = item[WikiSQLsql]
            query_index = set()
            query = set()
            query_index.add(sql[WikiSQLsel])
            for item in sql[WikiSQLconds]:
                query_index.add(item[0])
            for i in query_index:
                query.add(dbschema[dbid][Table][dbid][i])
            #query = set(question.lower().split(" "))
            questions[dbid].append((question, WikiSQL, query))
    elif dataname == Wikitable:
        for item in dataset:
            dbid = item[Wikitabledbid].lower()
            question = item[Question].lower()
            #query = set(question.lower().split(" "))
            query = set()
            questions[dbid].append((question, Wikitable, query))
        for item in datadb:
            table = datadb[item]
            dbschema[item.lower()][Table] = {item.lower(): table}
            dbschema[item.lower()][Column] = set(table)
    elif dataname == Kvret:
        for item in dataset:
            dbid = item[Kvretscenario][Kvretuuid].lower()
            for term in item[Kvretdialogue]:
                question = term[Kvretdata][Kvretutterance].lower()
                #query = set(question.lower().split(" "))
                query = set()
                questions[dbid].append((question, Kvret, query))
    elif dataname == Tablefact:
        for item in dataset:
            dbid = item
            for question in dataset[dbid][0]:
                #query = set(question.lower().split(" "))
                query = set()
                questions[dbid.split(".")[0]].append((question, Tablefact, query))
        for item in datadb:
            table = datadb[item]
            dbid = item.split(".")[0]
            dbschema[dbid][Table] = {dbid: table}
            dbschema[dbid][Column] = set(table)
    elif dataname == Msmarco:
        for item in dataset:
            dbid = Msmarco + str(len(questions))
            query = set()
            questions[dbid].append((item.lower(), Msmarco, query))
    elif dataname == WikiQA:
        for item in dataset:
            dbid = WikiQA + str(len(questions))
            query = set()
            questions[dbid].append((item.lower(), WikiQA, query))
    elif dataname == Coqa:
        coqadata = dataset[Coqadata]
        for item in coqadata:
            coqaquestions = item[Coqaquestions]
            for question in coqaquestions:
                dbid = Coqa + str(len(questions))
                query = set()
                questions[dbid].append((question[Coqainput].lower(), Coqa, query))
    elif dataname == Quac:
        quacdata = dataset[Quacdata]
        for item in quacdata:
            quacparagraphs = item[Quacparagraphs]
            for paragraph in quacparagraphs:
                quacqas = paragraph[Quacqas]
                for qa in quacqas:
                    question = qa[Quacquestion]
                    dbid = Quac + str(len(questions))
                    query = set()
                    questions[dbid].append((question.lower(), Quac, query))
    elif dataname == Dbdomainambiguous:
        for item in dataset:
            for question in dataset[item]:
                query = set()
                questions[item.lower()].append((question.lower(), Dbdomainambiguous, query))
        for item in datadb:
            tables = datadb[item]
            columns = set()
            dbschema[item][Table] = tables
            for table in tables:
                for column in tables[table]:
                    columns.add(column)
            dbschema[item][Column] = columns
    elif dataname == Dbdomainnotambiguous:
        for item in dataset:
            for question in dataset[item]:
                query = set()
                questions[item.lower()].append((question.lower(), Dbdomainnotambiguous, query))
        for item in datadb:
            tables = datadb[item]
            columns = set()
            dbschema[item][Table] = tables
            for table in tables:
                for column in tables[table]:
                    columns.add(column)
            dbschema[item][Column] = columns
    elif dataname == Alex:
        for item in dataset:
            question = item[Alexcontextuttl]
            query = set()
            dbid = Alex + str(len(questions))
            questions[dbid].append((question.lower(), Alex, query))
    elif dataname == Googlenq:
        for item in dataset:
            #question = item[Googlenqquestion]
            question = item
            query = set()
            dbid = Googlenq + str(len(questions))
            questions[dbid].append((question.lower(), Googlenq, query))
    elif dataname == Totto:
        for item in dataset:
            tottoquestions = item[Tottosentenceannotations]
            query = set()
            dbid = item[Tottotablesectiontitle]
            tottotables = item[Tottotable][0]
            for question in tottoquestions:
                questions[dbid].append((question[Tottofinalsentence].lower(), Totto, query))
            tables = {dbid: []}
            columns = set()
            for column in tottotables:
                tables[dbid].append(column[Tottovalue].lower())
                columns.add(column[Tottovalue].lower())
            dbschema[dbid][Table] = tables
            dbschema[dbid][Column] = columns
    elif dataname == Logicnlg:
        for item in dataset:
            dbid = item.split(".")[0]
            logicnlgquestions = dataset[item]
            query = set()
            for question in logicnlgquestions:
                questions[dbid].append((question[0].lower(), Logicnlg, query))
    elif dataname == Sparc:
        for item in dataset:
            dbid = item[Sparcdatabaseid].lower()
            interactions = item[Sparcinteraction]
            for interaction in interactions:
                query = set(interaction[Sparcquery].lower().split())
                question = interaction[Sparcutterance].lower()
                questions[dbid].append((question, Sparc, query))
            final = item[Sparcfinal]
            query = set(final[Sparcquery].lower().split())
            question = final[Sparcutterance]
            questions[dbid].append((question, Sparc, query))
        for item in datadb:
            dbid = item[Sparcdbid].lower()
            table_names = item[Sparctablename]
            table_names = dict(zip(range(len(table_names)), table_names))
            column_names = item[Sparccolumnname]
            table2column = defaultdict(list)
            column_set = set()
            for column in column_names:
                if column[0] in table_names:
                    table2column[table_names[column[0]].lower()].append(column[1].lower())
                column_set.add(column[1].lower())
            dbschema[dbid][Table] = table2column
            dbschema[dbid][Column] = column_set
    elif dataname == Cosql:
        for item in dataset:
            """dialog = dataset[item]
            dbid = dialog[Cosqldbid]
            question = dialog[Cosqlquerygoal].lower()
            query = set(dialog[Cosqlsql].lower().split(" "))
            questions[dbid].append((question, Cosql, query))"""

            dbid = item[Cosqluserdbid].lower()
            question = item[Cosqlutterance].lower()
            intent = item[Cosqlintent]
            query = set()
            if intent and intent[0] == Cosqlambiguous:
                questions[dbid].append((question, Cosql, query))
        for item in datadb:
            dbid = item[Cosqldbid].lower()
            table_names = item[Cosqltablename]
            table_names = dict(zip(range(len(table_names)), table_names))
            column_names = item[Cosqlcolumnname]
            table2column = defaultdict(list)
            column_set = set()
            for column in column_names:
                if column[0] in table_names:
                    table2column[table_names[column[0]].lower()].append(column[1].lower())
                column_set.add(column[1].lower())
            dbschema[dbid][Table] = table2column
            dbschema[dbid][Column] = column_set
    elif dataname == Cosqlnotambiguous:
        for item in dataset:
            dialog = dataset[item]
            dbid = dialog[Cosqldbid]
            question = dialog[Cosqlquerygoal].lower()
            query = set(dialog[Cosqlsql].lower().split(" "))
            questions[dbid].append((question, Cosql, query))
        for item in datadb:
            dbid = item[Cosqldbid].lower()
            table_names = item[Cosqltablename]
            table_names = dict(zip(range(len(table_names)), table_names))
            column_names = item[Cosqlcolumnname]
            table2column = defaultdict(list)
            column_set = set()
            for column in column_names:
                if column[0] in table_names:
                    table2column[table_names[column[0]].lower()].append(column[1].lower())
                column_set.add(column[1].lower())
            dbschema[dbid][Table] = table2column
            dbschema[dbid][Column] = column_set
    elif dataname == Alexa:
        for item in dataset:
            dbid = item
            datas = dataset[item]
            content = datas[Alexacontent]
            for item in content:
                question = item[Alexamessage]
                query = set()
                questions[dbid].append((question, Alexa, query))
    else:
        print("wrong type!")
        exit(0)

    return questions, dbschema

def splitdataschema(total_schema):
    print("schema_num: ")
    for datasetid in total_schema:
        print(datasetid + ": " + str(len(total_schema[datasetid])))
    print("------------------")

    train_schema = defaultdict(dict)
    dev_schema = defaultdict(dict)
    test_schema = defaultdict(dict)

    dbid2dataset = defaultdict(list)
    dbids = {}
    nonoverlapdbid = defaultdict(list)
    overlapdbid = defaultdict(list)
    tmpoverlapdbid = defaultdict(str)

    nonoverlapcolumns = []
    tmpcolumns = defaultdict(list)
    tmpcolumnsset = defaultdict(set)
    for datasetid in total_schema:
        dataset = total_schema[datasetid]
        for dbid in dataset:
            tables = total_schema[datasetid][dbid][Table]
            tmp = ""
            tablekeys = sorted(list(tables.keys()))
            for tableid in tablekeys:
                #tmp = str(sorted(tables[tableid]))
                tmp += '#'.join(sorted(tables[tableid])).lower().replace("_", " ") + "#"
            #.replace(" ", "").replace(".", "").replace("*", "")
            value = datasetid + "#" + dbid
            if not value in tmpcolumnsset[tmp]:
                tmpcolumns[tmp].append(value)
                tmpcolumnsset[tmp].add(value)

        dbids.update(dataset)

    for tmp in tmpcolumns:
        nonoverlapcolumns.append(tmpcolumns[tmp])

    overlapcount = defaultdict(int)

    for item in nonoverlapcolumns:
        if len(item) > 1:
            tmp = ""
            for term in item:
                tmpterm = term.split("#")[0]
                tmp += tmpterm + "-"
            tmp = "-".join(sorted(list(set(tmp[:-1].split("-")))))
            overlapcount[tmp] += 1

    print("overlap database columns:")

    for datasetid in overlapcount:
        print(datasetid + ": " + str(overlapcount[datasetid]))
    print("------------------")

    random.shuffle(nonoverlapcolumns)
    for i, item in enumerate(nonoverlapcolumns):
        overlaplen = len(nonoverlapcolumns)
        trainlen = int(Trainportion * overlaplen)
        devlen = int(Devportion * overlaplen)
        testlen = int(Testportion * overlaplen)
        for term in item:
            datasetid = term.split("#")[0]
            dbid = term.split("#")[1]

            if i < trainlen:
                train_schema[datasetid][dbid] = {'table': total_schema[datasetid][dbid][Table], 'column': total_schema[datasetid][dbid][Column]}
            elif i >= trainlen and i < devlen:
                dev_schema[datasetid][dbid] = {'table': total_schema[datasetid][dbid][Table], 'column': total_schema[datasetid][dbid][Column]}
            elif i >= devlen and i < testlen:
                test_schema[datasetid][dbid] = {'table': total_schema[datasetid][dbid][Table], 'column': total_schema[datasetid][dbid][Column]}

    """for datasetid in total_schema:
        dataset = total_schema[datasetid]
        for dbid in dataset:
            dbid2dataset[dbid].append(datasetid)
        dbids.update(dataset)

    for dbid in dbid2dataset:
        frq = len(dbid2dataset[dbid])
        datasetids = dbid2dataset[dbid]
        if frq > 1:
            tmpoverlapdbid[dbid] = "-".join(datasetids)
        else:
            nonoverlapdbid[datasetids[0]].append(dbid)

    overlapcount = defaultdict(int)
    for overlap in tmpoverlapdbid:
        overlapcount[tmpoverlapdbid[overlap]] += 1

    print("overlap database id:")
    for datasetid in overlapcount:
        print(datasetid + ": " + str(overlapcount[datasetid]))
    print("------------------")

    for dbid in tmpoverlapdbid:
        combineid = tmpoverlapdbid[dbid]
        overlapdbid[combineid].append(dbid)

    nonoverlapcolumns = []
    tmpcolumns = defaultdict(set)
    for datasetid in nonoverlapdbid:
        dataset = nonoverlapdbid[datasetid]
        for dbid in dataset:
            tables = total_schema[datasetid][dbid][Table]
            tmp = ""
            for tableid in tables:
                tmp += '#'.join(tables[tableid]) + "#"
            tmpcolumns[tmp].add(datasetid + "#" + dbid)

    for tmp in tmpcolumns:
        nonoverlapcolumns.append(tmpcolumns[tmp])

    overlapcount = defaultdict(int)

    for item in nonoverlapcolumns:
        if len(item) > 1:
            tmp = ""
            for term in item:
                tmpterm = term.split("#")[0]
                tmp += tmpterm + "-"
            tmp = tmp[:-1]
            overlapcount[tmp] += 1

    print("overlap database columns (in nonoverlap database id):")
    for datasetid in overlapcount:
        print(datasetid + ": " + str(overlapcount[datasetid]))
    print("------------------")

    for combineid in overlapdbid:
        overlapdbids = overlapdbid[combineid]
        random.shuffle(overlapdbids)
        overlaplen = len(overlapdbids)
        trainlen = int(Trainportion * overlaplen)
        devlen = int(Devportion * overlaplen)
        testlen = int(Testportion * overlaplen)
        datasetids = combineid.split("-")
        for datasetid in datasetids:
            for i, dbid in enumerate(overlapdbids):
                if i < trainlen:
                    train_schema[datasetid][dbid] = dbids[dbid]
                elif i >= trainlen and i < devlen:
                    dev_schema[datasetid][dbid] = dbids[dbid]
                elif i >= devlen and i < testlen:
                    test_schema[datasetid][dbid] = dbids[dbid]

    random.shuffle(nonoverlapcolumns)
    for i, item in enumerate(nonoverlapcolumns):
        overlaplen = len(nonoverlapcolumns)
        trainlen = int(Trainportion * overlaplen)
        devlen = int(Devportion * overlaplen)
        testlen = int(Testportion * overlaplen)
        for term in item:
            datasetid = term.split("#")[0]
            dbid = term.split("#")[1]
            if i < trainlen:
                train_schema[datasetid][dbid] = dbids[dbid]
            elif i >= trainlen and i < devlen:
                dev_schema[datasetid][dbid] = dbids[dbid]
            elif i >= devlen and i < testlen:
                test_schema[datasetid][dbid] = dbids[dbid]"""

    """for datasetid in nonoverlapdbid:
        nonoverlapdbids = nonoverlapdbid[datasetid]
        random.shuffle(nonoverlapdbids)
        nonoverlaplen = len(nonoverlapdbids)
        trainlen = int(Trainportion * nonoverlaplen)
        devlen = int(Devportion * nonoverlaplen)
        testlen = int(Testportion * nonoverlaplen)
        for i, dbid in enumerate(nonoverlapdbids):
            if i < trainlen:
                train_schema[datasetid][dbid] = dbids[dbid]
            elif i >= trainlen and i < devlen:
                dev_schema[datasetid][dbid] = dbids[dbid]
            elif i >= devlen and i < testlen:
                test_schema[datasetid][dbid] = dbids[dbid]"""


    print("train_schema_num:")
    for schema in train_schema:
        print(schema + ": " + str(len(train_schema[schema])))

    print("------------------")
    print("dev_schema_num:")
    for schema in dev_schema:
        print(schema + ": " + str(len(dev_schema[schema])))

    print("------------------")
    print("test_schema_num:")
    for schema in test_schema:
        print(schema + ": " + str(len(test_schema[schema])))

    print("------------------")


    return train_schema, dev_schema, test_schema

def filterquestion(totalq):
    total_question = defaultdict(lambda: defaultdict(list))
    questionsdict = {}

    for datasetid in totalq:
        dataset = totalq[datasetid]
        for dbid in dataset:
            database = dataset[dbid]
            for question, qdataset, query in database:
                lquestion = question.lower()
                if lquestion in questionsdict:
                    if query:
                        questionsdict[lquestion] = (dbid, qdataset, query)
                else:
                    questionsdict[lquestion] = (dbid, qdataset, query)

    print("question_num:")
    datasetquestion = defaultdict(int)
    for question in questionsdict:
        dbid = questionsdict[question][0]
        dataset = questionsdict[question][1]
        query = questionsdict[question][2]
        total_question[dataset][dbid].append((question, dataset, query))
        datasetquestion[dataset] += 1

    for dataset in datasetquestion:
        print(dataset + ": " + str(datasetquestion[dataset]))
    print("------------------")

    return total_question

def splittype1question(total_questions):
    tmpquestions = list()

    for datasetid in total_questions:
        if datasetid in type1qdataset:
            dataset = total_questions[datasetid]
            for dbid in dataset:
                questions = dataset[dbid]
                for question, qdataset, query in questions:
                    tmpquestions.append((qdataset, question))
    #questions = list(set(tmpquestions))

    questions = []
    questionset = set()
    for question in tmpquestions:
        if question not in questionset:
            questions.append(question)
            questionset.add(question)

    random.shuffle(questions)

    questionlen = len(questions)
    trainlen = int(Trainportion * questionlen)
    devlen = int(Devportion * questionlen)
    testlen = int(Testportion * questionlen)

    trainq = questions[: trainlen]
    devq = questions[trainlen: devlen]
    testq = questions[devlen: testlen]

    return trainq, devq, testq

def gen_type1(total_questions, total_schema, total_dataset, total_q):
    type1sample = []

    for i, (qdataset, question) in enumerate(total_questions):
        #randomkey1 = random.sample(type1dataset, 1)[0]
        randomkey1 = type1dataset[i % len(type1dataset)]
        databasekey = list(total_schema[randomkey1].keys())
        randomkey2 = random.sample(databasekey, 1)[0]
        schematable = total_schema[randomkey1][randomkey2][Table]
        if randomkey1 == Dbdomainnotambiguous or randomkey1 == Dbdomainambiguous:
            type1sample.append({Outtype: Type1, Outquestion: question, Outquestiondatasetid: qdataset,
                            Outdatabaseiddatasetid: randomkey2, Outdatabaseid: randomkey2, Outtables: schematable})
            #total_dataset[Type1][randomkey2] += 1
            #total_q[Type1][qdataset] += 1
        else:
            type1sample.append({Outtype: Type1, Outquestion: question, Outquestiondatasetid: qdataset,
                                Outdatabaseiddatasetid: randomkey1, Outdatabaseid: randomkey2, Outtables: schematable})
            #total_dataset[Type1][randomkey1] += 1
            #total_q[Type1][qdataset] += 1
    type1sample = random.sample(type1sample, int(len(type1sample) / 10))
    for item in type1sample:
        datasetid = item[Outdatabaseiddatasetid]
        qdataset = item[Outquestiondatasetid]
        total_dataset[Type1][datasetid] += 1
        total_q[Type1][qdataset] += 1

    print("type1: " + str(len(type1sample)))
    return type1sample

def gen_type2(total_questions, total_schema, total_dataset, total_q):
    type2sample = []

    for datasetid in total_schema:
        if datasetid in type2dataset:
            dataset = total_schema[datasetid]
            for dbid in dataset:
                database = dataset[dbid]
                schematable = database[Table]
                if dbid in total_questions[datasetid]:
                    for question, qdataset, query in total_questions[datasetid][dbid]:
                        if datasetid == Cosql:
                            type2sample.append({Outtype: Type2, Outquestion: question, Outquestiondatasetid: datasetid,
                                                Outdatabaseiddatasetid: datasetid, Outdatabaseid: dbid, Outtables: schematable})
                            total_dataset[Type2][datasetid] += 1
                            total_q[Type2][qdataset] += 1
                        else:
                            type2sample.append({Outtype: Type2, Outquestion: question, Outquestiondatasetid: dbid,
                                                Outdatabaseiddatasetid: dbid, Outdatabaseid: dbid,
                                                Outtables: schematable})
                            total_dataset[Type2][dbid] += 1
                            total_q[Type2][dbid] += 1

    print("type2: " + str(len(type2sample)))
    return type2sample

def gen_type3(total_questions, total_schema, total_dataset, total_q):
    type3sample = []
    deletenums = [1, 1, 2, 3]

    for datasetid in total_schema:
        if datasetid in type3dataset:
            dataset = total_schema[datasetid]
            for dbid in dataset:
                if dbid in total_questions[datasetid]:
                    database = dataset[dbid]
                    schematable = database[Table]
                    schemacolumn = database[Column]
                    questions = total_questions[datasetid][dbid]
                    for question, qdataset, query in questions:
                        if query & schemacolumn:
                            tmp_schematable = defaultdict(list)
                            deletecolumn = list(query & schemacolumn)
                            deletenum = random.sample(deletenums, 1)[0]
                            if deletenum > len(deletecolumn):
                                deletenum = len(deletecolumn)
                            deletecolumn = random.sample(deletecolumn, deletenum)
                            for key in schematable:
                                for item in schematable[key]:
                                    if item not in deletecolumn:
                                        tmp_schematable[key].append(item)
                            if datasetid == Dbdomainnotambiguous or datasetid == Spiderother:
                                type3sample.append({Outtype: Type3, Outquestion: question, Outquestiondatasetid: dbid,
                                                Outdatabaseiddatasetid: dbid, Outdatabaseid: dbid,
                                                Outtables: tmp_schematable})
                                total_dataset[Type3][dbid] += 1
                                total_q[Type3][dbid] += 1
                            elif datasetid == Cosqlnotambiguous:
                                type3sample.append(
                                    {Outtype: Type3, Outquestion: question, Outquestiondatasetid: Cosql,
                                     Outdatabaseiddatasetid: Cosql, Outdatabaseid: dbid,
                                     Outtables: tmp_schematable})
                                total_dataset[Type3][Cosql] += 1
                                total_q[Type3][Cosql] += 1
                            else:
                                type3sample.append(
                                    {Outtype: Type3, Outquestion: question, Outquestiondatasetid: datasetid,
                                     Outdatabaseiddatasetid: datasetid, Outdatabaseid: dbid,
                                     Outtables: tmp_schematable})
                                total_dataset[Type3][datasetid] += 1
                                total_q[Type3][datasetid] += 1
    print("type3: " + str(len(type3sample)))
    return type3sample

def gen_type4(total_questions, total_schema, total_dataset, total_q):
    type4sample = []

    for datasetid in total_schema:
        if datasetid in type4dataset:
            dataset = total_schema[datasetid]
            for dbid in dataset:
                database = dataset[dbid]
                schematable = database[Table]
                if dbid in total_questions[datasetid]:
                    for question, qdataset, query in total_questions[datasetid][dbid]:
                        type4sample.append({Outtype: Type4, Outquestion: question, Outquestiondatasetid: datasetid,
                                            Outdatabaseiddatasetid: datasetid, Outdatabaseid: dbid,
                                            Outtables: schematable})
                        total_dataset[Type4][datasetid] += 1
                        total_q[Type4][datasetid] += 1

    print("type4: " + str(len(type4sample)))
    return type4sample

def gen_type5(total_questions, total_schema, total_dataset, total_q):
    type5sample = []
    deletenums = [1, 1, 2, 3]

    for datasetid in total_schema:
        if datasetid in type5dataset:
            dataset = total_schema[datasetid]
            for dbid in dataset:
                database = dataset[dbid]
                schematable = database[Table]
                schemacolumn = database[Column]
                if dbid in total_questions[datasetid]:
                    for question, qdataset, query in total_questions[datasetid][dbid]:
                        if datasetid == Dbdomainnotambiguous or datasetid == Spiderother:
                            type5sample.append({Outtype: Type5, Outquestion: question, Outquestiondatasetid: dbid,
                                            Outdatabaseiddatasetid: dbid, Outdatabaseid: dbid,
                                            Outtables: schematable})
                            total_dataset[Type5][dbid] += 1
                            total_q[Type5][dbid] += 1
                        elif datasetid == Cosqlnotambiguous:
                            type5sample.append({Outtype: Type5, Outquestion: question, Outquestiondatasetid: Cosql,
                                                Outdatabaseiddatasetid: Cosql, Outdatabaseid: dbid,
                                                Outtables: schematable})
                            total_dataset[Type5][Cosql] += 1
                            total_q[Type5][Cosql] += 1
                        else:
                            type5sample.append({Outtype: Type5, Outquestion: question, Outquestiondatasetid: datasetid,
                                                Outdatabaseiddatasetid: datasetid, Outdatabaseid: dbid,
                                                Outtables: schematable})
                            total_dataset[Type5][datasetid] += 1
                            total_q[Type5][datasetid] += 1
                        tmp_schematable = defaultdict(list)
                        deletecolumn = list(schemacolumn - query)
                        deletenum = random.sample(deletenums, 1)[0]
                        if deletenum > len(deletecolumn):
                            deletenum = len(deletecolumn)
                        deletecolumn = random.sample(deletecolumn, deletenum)
                        for key in schematable:
                            for item in schematable[key]:
                                if item not in deletecolumn:
                                    tmp_schematable[key].append(item)
                        if datasetid == Dbdomainnotambiguous or datasetid == Spiderother:
                            type5sample.append({Outtype: Type5, Outquestion: question, Outquestiondatasetid: dbid,
                                            Outdatabaseiddatasetid: dbid, Outdatabaseid: dbid,
                                            Outtables: tmp_schematable})
                            total_dataset[Type5][dbid] += 1
                            total_q[Type5][dbid] += 1
                        elif datasetid == Cosqlnotambiguous:
                            type5sample.append({Outtype: Type5, Outquestion: question, Outquestiondatasetid: Cosql,
                                                Outdatabaseiddatasetid: Cosql, Outdatabaseid: dbid,
                                                Outtables: tmp_schematable})
                            total_dataset[Type5][Cosql] += 1
                            total_q[Type5][Cosql] += 1
                        else:
                            type5sample.append({Outtype: Type5, Outquestion: question, Outquestiondatasetid: datasetid,
                                                Outdatabaseiddatasetid: datasetid, Outdatabaseid: dbid,
                                                Outtables: tmp_schematable})
                            total_dataset[Type5][datasetid] += 1
                            total_q[Type5][datasetid] += 1
    print("type5: " + str(len(type5sample)))
    return type5sample

def write_json(samples, path):
    f = open(path, "w", encoding="utf-8")
    json.dump(samples, f, sort_keys=True, indent=4, separators=(',', ': '))
    f.close()

def draw_stat(dataset, name):
    types = [Type1, Type2, Type3, Type4, Type5]
    datasetname = set()
    for t in dataset:
        datasetname |= set(dataset[t].keys())
    datasetname = list(datasetname)
    datasetstats = []
    for n in datasetname:
        datasetstat = []
        for t in types:
            datasetstat.append(dataset[t][n])
        datasetstats.append(datasetstat)

    bar = Bar("")
    for index, n in enumerate(datasetname):
        bar.add(n, types, datasetstats[index], is_stack=True, is_more_utils=True)
    bar.render(name + '_stat_graph.html')

def get_stat(train_questions, test_questions, dev_questions):
    group1 = [Spider, Sparc, Cosql]
    group2 = [WikiSQL]
    group3 = ['restaurants', 'scholar', 'yelp', 'imdb', 'geo', 'academic']
    group4 = [Tablefact, Totto, Logicnlg]
    group5 = [HybridQA]
    group6 = [Wikitable]
    group7 = [Alexa, Googlenq, Msmarco, WikiQA, Coqa, Quac]
    group1dict = defaultdict(int)
    group2dict = defaultdict(int)
    group3dict = defaultdict(int)
    group4dict = defaultdict(int)
    group5dict = defaultdict(int)
    group6dict = defaultdict(int)
    group7dict = defaultdict(int)

    for t in train_questions:
        datasets = train_questions[t]
        for dataset in datasets:
            if dataset in group1:
                group1dict[t] += datasets[dataset]
            elif dataset in group2:
                group2dict[t] += datasets[dataset]
            elif dataset in group3:
                group3dict[t] += datasets[dataset]
            elif dataset in group4:
                group4dict[t] += datasets[dataset]
            elif dataset in group5:
                group5dict[t] += datasets[dataset]
            elif dataset in group6:
                group6dict[t] += datasets[dataset]
            elif dataset in group7:
                group7dict[t] += datasets[dataset]
            else:
                print(dataset)
                print("train wrong!")

    for t in test_questions:
        datasets = test_questions[t]
        for dataset in datasets:
            if dataset in group1:
                group1dict[t] += datasets[dataset]
            elif dataset in group2:
                group2dict[t] += datasets[dataset]
            elif dataset in group3:
                group3dict[t] += datasets[dataset]
            elif dataset in group4:
                group4dict[t] += datasets[dataset]
            elif dataset in group5:
                group5dict[t] += datasets[dataset]
            elif dataset in group6:
                group6dict[t] += datasets[dataset]
            elif dataset in group7:
                group7dict[t] += datasets[dataset]
            else:
                print(dataset)
                print("test wrong!")

    for t in dev_questions:
        datasets = dev_questions[t]
        for dataset in datasets:
            if dataset in group1:
                group1dict[t] += datasets[dataset]
            elif dataset in group2:
                group2dict[t] += datasets[dataset]
            elif dataset in group3:
                group3dict[t] += datasets[dataset]
            elif dataset in group4:
                group4dict[t] += datasets[dataset]
            elif dataset in group5:
                group5dict[t] += datasets[dataset]
            elif dataset in group6:
                group6dict[t] += datasets[dataset]
            elif dataset in group7:
                group7dict[t] += datasets[dataset]
            else:
                print(dataset)
                print("dev wrong!")

    print("group1: " + str(dict(group1dict)))
    print("group2: " + str(dict(group2dict)))
    print("group3: " + str(dict(group3dict)))
    print("group4: " + str(dict(group4dict)))
    print("group5: " + str(dict(group5dict)))
    print("group6: " + str(dict(group6dict)))
    print("group7: " + str(dict(group7dict)))

if __name__ == "__main__":
    from openpyxl import load_workbook
    wb = load_workbook('testset.xlsx')
    sheets = wb.get_sheet_names()
    datasets = defaultdict(lambda :defaultdict(int))
    for sheet in sheets:
        table = wb.get_sheet_by_name(sheet)
        rows = table.max_row
        for row in range(rows):
            dataset = table.cell(row=row + 1, column=1).value
            num = table.cell(row=row + 1, column=2).value
            datasets[sheet][dataset] = num
    draw_stat(datasets, "annotate")


